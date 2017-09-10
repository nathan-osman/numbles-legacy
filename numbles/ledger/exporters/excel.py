from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .exporter import Exporter


class ExcelExporter(Exporter):
    """
    Create an Excel 2010 file
    """

    def export(self, transactions):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="numbles.xlsx"'
        wb = Workbook()
        ws = wb.active
        font = Font(bold=True)
        for i, c in zip(range(0, len(self.COLUMNS)), self.COLUMNS):
            cell = ws.cell(row=1, column=(i + 1))
            cell.value = self.TITLES[c]
            cell.font = font
        for j, t in zip(range(0, transactions.count()), transactions):
            for i, c in zip(range(0, len(self.COLUMNS)), self.COLUMNS):
                cell = ws.cell(row=(j + 2), column=(i + 1))
                cell.value = self.column(c, t)
        wb.save(response)
        return response
